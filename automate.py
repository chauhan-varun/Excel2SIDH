"""
Skill India Digital - Automated Marks Entry
============================================
Reads student marks from Excel and enters them on the SIDH portal
using Playwright browser automation.

Flow:
  1. Opens browser → YOU log in manually
  2. Press ENTER in terminal when ready
  3. Script takes over the SAME browser session and fills all marks

Usage:
    uv run python automate.py                    # Normal run (you log in first)
    uv run python automate.py --dry-run          # Fill marks but don't submit
    uv run python automate.py --start-from 5     # Start from student #5
    uv run python automate.py --student CAN_XXX  # Single student only
"""

import argparse
import json
import logging
import os
import re
import sys
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl
from playwright.sync_api import sync_playwright, Page, BrowserContext, Locator

# ─── Configuration ───────────────────────────────────────────────────────────

BATCH_ID = "3391657"
BASE_URL = "https://admin.skillindiadigital.gov.in"
BATCH_PAGE_URL = f"{BASE_URL}/admin-profile/assessor/master-assessor/view-batch-details-new/PENDING/{BATCH_ID};batches=assessment"
EXCEL_FILE = "Result Sheet _ 3391657, 21-05-2026.xlsx"
SHEET_NAME = "Detailed Result "

# Timing (seconds) - adjust if portal is slow
DELAY_BETWEEN_FIELDS = 0.15       # delay between typing each field
DELAY_AFTER_SEARCH = 2.0          # wait after searching a student
DELAY_AFTER_NAVIGATE = 2.5        # wait after page navigation
DELAY_BEFORE_SUBMIT = 1.0         # wait before clicking submit
DELAY_AFTER_SUBMIT = 2.0          # wait after submitting marks

# Excel column mapping
STUDENT_DATA_START_ROW = 4        # first student row
ENROLLMENT_COL = 5                # column E
NAME_COL = 8                      # column H
EXCEL_GRAND_TOTAL_COL = 85        # column CG ("Grand Total-( Theory, Skills Practical and Viva)")

# NOS mapping: (NOS code, start_column_in_excel)
# Each NOS has 4 columns: Theory, Practical, OJT, Viva
NOS_COLUMNS = [
    ("HSS/N5125", 13),
    ("HSS/N5126", 17),
    ("HSS/N5106", 21),
    ("HSS/N5112", 25),
    ("HSS/N5113", 29),
    ("HSS/N5127", 33),
    ("HSS/N5115", 37),
    ("HSS/N5128", 41),
    ("HSS/N9615", 45),
    ("HSS/N9616", 49),
    ("HSS/N9617", 53),
    ("HSS/N9618", 57),
    ("DGT/VSQ/N0102", 61),
    ("HSS/N5129(Core)", 65),
    ("HSS/N5130(Core)", 69),
    ("HSS/N5131(Core)", 73),
    ("HSS/N5132(Core)", 77),
]

# ─── Logging ─────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)

# ─── Data Classes ────────────────────────────────────────────────────────────

@dataclass
class NOSMarks:
    """Marks for a single NOS code"""
    nos_code: str
    theory: float
    practical: float
    ojt: float
    viva: float

    @property
    def total(self) -> float:
        return self.theory + self.practical + self.ojt + self.viva


@dataclass
class StudentData:
    """All data for a single student"""
    serial_no: int
    enrollment_no: str
    name: str
    excel_grand_total: float = 0.0
    nos_marks: list[NOSMarks] = field(default_factory=list)

    @property
    def grand_total(self) -> float:
        return self.excel_grand_total


# ─── Excel Reader ────────────────────────────────────────────────────────────

def read_excel_data(filepath: str) -> list[StudentData]:
    """Read all student marks data from the Excel file."""
    log.info(f"Reading Excel file: {filepath}")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[SHEET_NAME]

    students = []
    row = STUDENT_DATA_START_ROW

    while row <= ws.max_row:
        enrollment = ws.cell(row, ENROLLMENT_COL).value
        if not enrollment:
            row += 1
            continue

        name = ws.cell(row, NAME_COL).value or "Unknown"
        serial = ws.cell(row, 1).value or (len(students) + 1)

        excel_total = ws.cell(row, EXCEL_GRAND_TOTAL_COL).value
        student = StudentData(
            serial_no=int(serial),
            enrollment_no=str(enrollment).strip(),
            name=str(name).strip(),
            excel_grand_total=float(excel_total or 0.0),
        )

        # Read marks for each NOS
        for nos_code, start_col in NOS_COLUMNS:
            theory = ws.cell(row, start_col).value
            practical = ws.cell(row, start_col + 1).value
            ojt = ws.cell(row, start_col + 2).value
            viva = ws.cell(row, start_col + 3).value

            student.nos_marks.append(NOSMarks(
                nos_code=nos_code,
                theory=float(theory or 0),
                practical=float(practical or 0),
                ojt=float(ojt or 0),
                viva=float(viva or 0),
            ))

        students.append(student)
        row += 1

    log.info(f"Loaded {len(students)} students from Excel")
    return students


# ─── Browser Automation ──────────────────────────────────────────────────────

class SIDHAutomation:
    """Automates marks entry on Skill India Digital Hub portal."""

    def __init__(self, page: Page, dry_run: bool = False):
        self.page = page
        self.dry_run = dry_run
        self.results: list[dict] = []

    def wait_for_load(self, timeout: float = DELAY_AFTER_NAVIGATE):
        """Wait for page to be ready."""
        time.sleep(timeout)
        try:
            self.page.wait_for_load_state("networkidle", timeout=10000)
        except Exception:
            pass  # timeout is ok, page might have polling requests

    def wait_for_any_locator(self, locators: list[str], timeout_ms: float = 5000) -> Optional[Locator]:
        """Wait for any of the given locator strings to become visible and enabled."""
        start = time.time()
        while (time.time() - start) * 1000 < timeout_ms:
            for loc_str in locators:
                try:
                    candidates = self.page.locator(loc_str).all()
                    for cand in candidates:
                        if cand.is_visible() and cand.is_enabled():
                            return cand
                except Exception:
                    pass
            time.sleep(0.2)
        return None

    def navigate_to_batch(self):
        """Navigate to the batch details page if not already there."""
        if BATCH_PAGE_URL not in self.page.url:
            log.info(f"Navigating to batch page: {BATCH_PAGE_URL}")
            self.page.goto(BATCH_PAGE_URL, wait_until="domcontentloaded")
            self.wait_for_load()
        else:
            log.info("Already on the batch page")
        log.info(f"Page title: {self.page.title()}")

    def click_approved_applicants_tab(self):
        """Click on the 'Approved Applicant' tab."""
        log.info("Looking for 'Approved Applicant' tab...")

        selectors = [
            "text=Approved Applicants",
            "a:has-text('Approved Applicants')",
            "text=Approved Applicant",
            "text=Approved applicant",
            "text=APPROVED APPLICANT",
            "[role='tab']:has-text('Approved')",
            "a:has-text('Approved')",
            "button:has-text('Approved')",
            ".mat-tab-label:has-text('Approved')",
        ]

        el = self.wait_for_any_locator(selectors, timeout_ms=8000)
        if el:
            try:
                el.click()
                log.info("Clicked 'Approved Applicant' tab")
                self.wait_for_load()
                return True
            except Exception as e:
                log.error(f"Failed to click Approved Applicants tab: {e}")

        log.warning("Could not find 'Approved Applicant' tab - it might already be selected")
        return False

    def search_student(self, enrollment_no: str) -> bool:
        """Search for a student by enrollment number."""
        log.info(f"Searching for student: {enrollment_no}")

        search_selectors = [
            "input[placeholder*='Search by Applicants']",
            "input[placeholder*='Applicants Name']",
            "input[placeholder*='Search']",
            "input[placeholder*='search']",
            "input[placeholder*='Applicant']",
            "input[type='text']",
            "input[type='search']",
        ]

        search_input = self.wait_for_any_locator(search_selectors, timeout_ms=8000)

        if not search_input:
            log.error("Could not find search input field!")
            return False

        # Clear and type enrollment number
        search_input.click()
        search_input.fill("")
        time.sleep(0.1)
        search_input.type(enrollment_no, delay=30)
        time.sleep(0.2)
        search_input.press("Enter")

        # Click the search button icon if present (ID: basic-addon2)
        try:
            search_btn = self.page.locator("button#basic-addon2").first
            if search_btn.is_visible(timeout=1000):
                search_btn.click()
                log.info("Clicked search button")
        except Exception:
            pass

        time.sleep(DELAY_AFTER_SEARCH)

        # Check if student appeared in results
        try:
            student_row = self.page.locator(f"text={enrollment_no}").first
            if student_row.is_visible(timeout=5000):
                log.info(f"Student found: {enrollment_no}")
                return True
        except Exception:
            pass

        log.warning(f"Student {enrollment_no} might not be visible in search results")
        return True  # Continue anyway

    def click_view_job_role(self, enrollment_no: str) -> bool:
        """Click the three-dot menu and select 'View Job Role Detail' or navigate directly."""
        log.info(f"Opening job role details for {enrollment_no}...")

        # 1. Try direct link extraction first (much faster and more robust)
        try:
            row_locator = self.page.locator(f"tr:has-text('{enrollment_no}')").first
            if row_locator.is_visible(timeout=2000):
                # Search for any link in this row containing the enrollment number
                link = row_locator.locator(f"a[href*='{enrollment_no}']").first
                if not link.is_visible(timeout=1000):
                    # Try general class for dropdown items in this row
                    link = row_locator.locator("a.dropdown-item").first
                
                if link.is_visible(timeout=1000):
                    href = link.get_attribute("href")
                    if href:
                        target_url = href if href.startswith("http") else f"{BASE_URL.rstrip('/')}/{href.lstrip('/')}"
                        log.info(f"Found direct link: {target_url}. Navigating directly...")
                        self.page.goto(target_url, wait_until="domcontentloaded")
                        self.wait_for_load()
                        return True
        except Exception as e:
            log.warning(f"Direct link navigation failed: {e}. Falling back to click path.")

        # 2. Click path fallback
        try:
            # Find the row containing this enrollment number, then the menu button
            row_selectors = [
                f"tr:has-text('{enrollment_no}')",
                f"mat-row:has-text('{enrollment_no}')",
                f"[class*='row']:has-text('{enrollment_no}')",
                f"div:has-text('{enrollment_no}')",
            ]

            for row_sel in row_selectors:
                try:
                    row = self.page.locator(row_sel).first
                    if not row.is_visible(timeout=2000):
                        continue

                    # Find three-dot menu button in this row
                    menu_selectors = [
                        "button:has(mat-icon:has-text('more_vert'))",
                        "button mat-icon:has-text('more_vert')",
                        "[matMenuTriggerFor]",
                        "button:has(.mat-icon)",
                        ".three-dot",
                        "mat-icon:has-text('more')",
                        "button.mat-icon-button",
                        "a[data-toggle='dropdown']",
                        ".dropdown a",
                    ]

                    for menu_sel in menu_selectors:
                        try:
                            menu_btn = row.locator(menu_sel).first
                            if menu_btn.is_visible(timeout=1000):
                                menu_btn.click()
                                time.sleep(0.5)
                                log.info("Clicked three-dot menu")
                                break
                        except Exception:
                            continue
                    else:
                        continue
                    break
                except Exception:
                    continue

            # Click "View Job Role Detail" from the dropdown
            time.sleep(0.5)
            view_selectors = [
                "text=View Job Role Details",
                "text=view job role details",
                "text=View Job Role Detail",
                "text=view job role detail",
                "text=View job role",
                "button:has-text('View Job Role')",
                "[role='menuitem']:has-text('View')",
                "mat-menu-item:has-text('View')",
                ".mat-menu-item:has-text('View')",
                "a.dropdown-item",
            ]

            for sel in view_selectors:
                try:
                    el = self.page.locator(sel).first
                    if el.is_visible(timeout=2000):
                        el.click()
                        log.info("Clicked 'View Job Role Detail'")
                        self.wait_for_load()
                        return True
                except Exception:
                    continue

            log.error("Could not find 'View Job Role Detail' menu item")
            return False

        except Exception as e:
            log.error(f"Error clicking view job role: {e}")
            return False

    def check_marks_already_uploaded(self) -> bool:
        """Check if marks have already been uploaded for this student."""
        try:
            # Using exact regex match to avoid matching "Not Uploaded"
            uploaded_selectors = [
                r"text=/^Uploaded$/i",
                r"text=/^\s*Uploaded\s*$/i",
                r"text=/^Already uploaded$/i",
                r"text=/^\s*Already uploaded\s*$/i",
                ".status-uploaded",
            ]

            for sel in uploaded_selectors:
                try:
                    el = self.page.locator(sel).first
                    if el.is_visible(timeout=1000):
                        log.info(f"Detected already uploaded status via selector: {sel}")
                        return True
                except Exception:
                    continue

            return False
        except Exception:
            return False

    def click_upload_marks_on_job_role_page(self) -> bool:
        """Click the three-dot menu on the job role details page and select 'Upload Marks'."""
        log.info("Looking for 'Upload Marks' button on the job role page...")
        try:
            menu_selectors = [
                "tr button:has(mat-icon:has-text('more_vert'))",
                "tr button mat-icon:has-text('more_vert')",
                "tr .la-ellipsis-h",
                "tr [data-toggle='dropdown']",
                "tr a[data-toggle='dropdown']",
                "tr .dropdown a",
                "button:has(mat-icon:has-text('more_vert'))",
                "[data-toggle='dropdown']",
            ]

            menu_btn = self.wait_for_any_locator(menu_selectors, timeout_ms=8000)
            if not menu_btn:
                log.error("Could not find three-dot menu on job role page")
                return False

            menu_btn.click()
            time.sleep(0.5)
            log.info("Clicked job role three-dot menu")

            upload_selectors = [
                "text=Upload Marks",
                "text=upload marks",
                "text=Upload",
                "a:has-text('Upload Marks')",
                "button:has-text('Upload Marks')",
                ".dropdown-item:has-text('Upload')",
            ]

            upload_btn = self.wait_for_any_locator(upload_selectors, timeout_ms=4000)
            if not upload_btn:
                log.error("Could not find 'Upload Marks' menu item")
                return False

            upload_btn.click()
            self.wait_for_load()
            log.info("Clicked 'Upload Marks'")
            return True

        except Exception as e:
            log.error(f"Error clicking Upload Marks: {e}")
            return False

    def find_nos_row(self, nos_code: str) -> Optional[Locator]:
        """Find the form row for a specific NOS code."""
        clean_code = nos_code.replace("(Core)", "").strip()

        selectors = [
            f"tr:has-text('{nos_code}')",
            f"tr:has-text('{clean_code}')",
            f"div:has-text('{nos_code}')",
            f"[class*='row']:has-text('{nos_code}')",
            f"mat-row:has-text('{nos_code}')",
        ]

        for sel in selectors:
            try:
                row = self.page.locator(sel).first
                if row.is_visible(timeout=1000):
                    return row
            except Exception:
                continue

        return None

    def _fill_nos_marks_fallback(self, nos_row: Locator, nos_marks: NOSMarks) -> bool:
        """Fallback method to fill marks using sequential input order."""
        try:
            inputs = nos_row.locator("input[type='number'], input[type='text'], input").all()
            if len(inputs) == 0:
                log.error(f"No input fields found for {nos_marks.nos_code} in fallback!")
                return False

            marks_values = [
                ("Theory", nos_marks.theory),
                ("Practical", nos_marks.practical),
                ("OJT", nos_marks.ojt),
                ("Viva", nos_marks.viva),
            ]

            for idx, (mark_type, value) in enumerate(marks_values):
                if idx < len(inputs):
                    inp = inputs[idx]
                    str_value = str(int(value)) if value == int(value) else str(value)
                    inp.click()
                    inp.fill("")
                    time.sleep(0.05)
                    inp.fill(str_value)
                    time.sleep(DELAY_BETWEEN_FIELDS)
            return True
        except Exception as e:
            log.error(f"Error in fallback filling: {e}")
            return False

    def fill_nos_marks(self, nos_marks: NOSMarks) -> bool:
        """Fill marks for a single NOS code mapping inputs by table column index."""
        nos_row = self.find_nos_row(nos_marks.nos_code)

        if not nos_row:
            log.warning(f"Could not find NOS row for: {nos_marks.nos_code}")
            return False

        try:
            cells = nos_row.locator("td").all()
            
            # Determine column mapping dynamically based on whether it is an Elective row or has 6/7 cells:
            # Elective rows (containing "(Core)") lack the "Type" cell, even if len(cells) evaluates to 7 due to hidden layout elements.
            is_elective = "(Core)" in nos_marks.nos_code
            try:
                row_text = nos_row.text_content() or ""
                if "elective" in row_text.lower():
                    is_elective = True
            except Exception:
                pass

            if is_elective:
                mappings = [
                    ("Theory", 1, nos_marks.theory),
                    ("Practical", 2, nos_marks.practical),
                    ("OJT", 3, nos_marks.ojt),
                    ("Viva", 4, nos_marks.viva),
                ]
            elif len(cells) == 7:
                mappings = [
                    ("Theory", 2, nos_marks.theory),
                    ("Practical", 3, nos_marks.practical),
                    ("OJT", 4, nos_marks.ojt),
                    ("Viva", 5, nos_marks.viva),
                ]
            elif len(cells) == 6:
                mappings = [
                    ("Theory", 1, nos_marks.theory),
                    ("Practical", 2, nos_marks.practical),
                    ("OJT", 3, nos_marks.ojt),
                    ("Viva", 4, nos_marks.viva),
                ]
            else:
                log.warning(f"Unexpected number of cells ({len(cells)}) in NOS row. Using fallback.")
                return self._fill_nos_marks_fallback(nos_row, nos_marks)

            filled_fields = []
            for mark_type, cell_idx, value in mappings:
                if cell_idx < len(cells):
                    cell = cells[cell_idx]
                    inp = cell.locator("input").first
                    if inp.is_visible(timeout=500) and inp.is_enabled(timeout=500):
                        # Convert to int if whole number, else keep decimal
                        str_value = str(int(value)) if value == int(value) else str(value)

                        inp.click()
                        inp.fill("")
                        time.sleep(0.05)
                        inp.type(str_value, delay=50)
                        time.sleep(DELAY_BETWEEN_FIELDS)

                        # Verify the value was entered correctly
                        entered = inp.input_value()
                        if entered != str_value:
                            log.warning(f"  Mismatch in {nos_marks.nos_code} {mark_type}: "
                                        f"expected '{str_value}', got '{entered}'")
                        else:
                            log.debug(f"  {nos_marks.nos_code} {mark_type}: {str_value} ✓")
                            filled_fields.append(mark_type)

            log.info(f"  ✓ {nos_marks.nos_code}: T={nos_marks.theory} P={nos_marks.practical} "
                     f"O={nos_marks.ojt} V={nos_marks.viva} (Filled: {', '.join(filled_fields) if filled_fields else 'None'})")
            
            # Add a small delay between rows for stability
            time.sleep(0.8)
            return True

        except Exception as e:
            log.error(f"Error filling marks for {nos_marks.nos_code}: {e}")
            return False

    def verify_grand_total(self, expected_total: float) -> bool:
        """Verify the form's calculated grand total matches the Excel total."""
        try:
            total_selectors = [
                "text=/Grand Total/i",
                "text=/Total Marks/i",
                "td:has-text('Total')",
                ".grand-total",
                "tfoot",
            ]

            for sel in total_selectors:
                try:
                    el = self.page.locator(sel).first
                    if el.is_visible(timeout=2000):
                        text = el.text_content() or ""
                        # Parse obtained total: look for a number followed by '/' (e.g. '1031.75/1300')
                        match = re.search(r'(\d+\.?\d*)\s*/', text)
                        if match:
                            form_total = float(match.group(1))
                        else:
                            # Fallback: take the first number in the text
                            numbers = re.findall(r'\d+\.?\d*', text)
                            if numbers:
                                form_total = float(numbers[0])
                            else:
                                continue

                        if abs(form_total - expected_total) < 1:
                            log.info(f"  ✓ Grand total verified: {form_total} (expected: {expected_total})")
                            return True
                        else:
                            log.warning(f"  ⚠ Grand total mismatch: form={form_total}, excel={expected_total}")
                            return False
                except Exception:
                    continue

            log.warning("Could not find grand total on page to verify")
            return True

        except Exception as e:
            log.error(f"Error verifying grand total: {e}")
            return True

    def submit_marks(self) -> bool:
        """Click the submit/upload button to save marks."""
        if self.dry_run:
            log.info("  [DRY RUN] Would submit marks here - skipping")
            return True

        time.sleep(DELAY_BEFORE_SUBMIT)

        submit_selectors = [
            "button:has-text('SAVE & UPLOAD')",
            "button:has-text('Save & Upload')",
            "button:has-text('save & upload')",
            "button:has-text('Upload')",
            "button:has-text('upload')",
            "button:has-text('Submit')",
            "button:has-text('submit')",
            "button:has-text('Save')",
            "button:has-text('save')",
            "button[type='submit']",
            "input[type='submit']",
        ]

        for sel in submit_selectors:
            try:
                btn = self.page.locator(sel).first
                if btn.is_visible(timeout=2000):
                    btn.click()
                    log.info("  Clicked submit/upload button. Waiting for confirmation popup...")
                    
                    # Wait for and click the 'OK' confirmation button in the popup
                    ok_selectors = [
                        "button:has-text('OK')",
                        "button:has-text('ok')",
                        "button:has-text('Ok')",
                        ".swal2-confirm",
                        ".confirm",
                        "button.btn-primary:has-text('OK')",
                    ]
                    ok_btn = self.wait_for_any_locator(ok_selectors, timeout_ms=8000)
                    if ok_btn:
                        ok_btn.click()
                        log.info("  ✓ Clicked OK on success popup")
                    else:
                        log.warning("  ⚠ Did not detect success popup OK button, continuing anyway")
                    
                    time.sleep(DELAY_AFTER_SUBMIT)
                    log.info("  ✓ Marks submitted/uploaded")
                    return True
            except Exception:
                continue

        log.error("  ✗ Could not find submit/upload button!")
        return False

    def go_back_to_batch_list(self):
        """Navigate back to the batch details / approved applicants list."""
        back_selectors = [
            "text=View Batches",
            "text=view batches",
            "a:has-text('View Batch')",
            "text=Back",
            "button:has-text('Back')",
        ]

        for sel in back_selectors:
            try:
                el = self.page.locator(sel).first
                if el.is_visible(timeout=2000):
                    el.click()
                    self.wait_for_load()
                    log.info("Navigated back to batch list")
                    return
            except Exception:
                continue

        # Fallback: navigate directly
        log.info("Using direct navigation to go back to batch page")
        self.page.goto(BATCH_PAGE_URL, wait_until="domcontentloaded")
        self.wait_for_load()
        self.click_approved_applicants_tab()

    def process_student(self, student: StudentData, total_students: int) -> dict:
        """Process a single student - search, fill marks, submit."""
        result = {
            "serial": student.serial_no,
            "enrollment": student.enrollment_no,
            "name": student.name,
            "status": "pending",
            "error": None,
        }

        try:
            log.info(f"\n{'='*60}")
            log.info(f"Processing student {student.serial_no}/{total_students}: "
                     f"{student.name} ({student.enrollment_no})")
            log.info(f"{'='*60}")

            # Step 1: Search for student
            if not self.search_student(student.enrollment_no):
                result["status"] = "search_failed"
                result["error"] = "Could not find student in search"
                return result

            # Step 2: Click three-dot menu → View Job Role Detail
            if not self.click_view_job_role(student.enrollment_no):
                result["status"] = "navigation_failed"
                result["error"] = "Could not open job role details"
                return result

            # Step 3: Check if marks already uploaded
            if self.check_marks_already_uploaded():
                log.info(f"  ℹ Marks already uploaded for {student.enrollment_no} - skipping")
                result["status"] = "already_uploaded"
                self.go_back_to_batch_list()
                return result

            # Step 3.5: Click three dots -> Upload Marks on the job role details page
            if not self.click_upload_marks_on_job_role_page():
                result["status"] = "upload_marks_navigation_failed"
                result["error"] = "Could not open upload marks form"
                self.go_back_to_batch_list()
                return result

            # Wait for the first NOS row to load in the DOM
            if student.nos_marks:
                first_nos = student.nos_marks[0].nos_code
                log.info(f"Waiting for form to load (looking for {first_nos})...")
                try:
                    self.page.locator(f"tr:has-text('{first_nos}')").first.wait_for(state="visible", timeout=12000)
                    log.info("Form loaded successfully.")
                except Exception as e:
                    log.warning(f"Timeout waiting for form to load: {e}")

            # Step 4: Fill marks for each NOS
            filled_count = 0
            for nos_marks in student.nos_marks:
                if self.fill_nos_marks(nos_marks):
                    filled_count += 1

            log.info(f"  Filled {filled_count}/{len(student.nos_marks)} NOS codes")

            if filled_count == 0:
                result["status"] = "no_fields_found"
                result["error"] = "Could not find any NOS input fields"
                self.go_back_to_batch_list()
                return result

            # Step 5: Verify grand total
            if not self.verify_grand_total(student.grand_total):
                result["status"] = "total_mismatch"
                result["error"] = f"Grand total mismatch: expected {student.grand_total}"
                log.error(f"  ✗ Grand total verification failed! Skipping submit to prevent incorrect upload.")
                self.go_back_to_batch_list()
                return result

            # Step 6: Submit marks
            if self.submit_marks():
                result["status"] = "success"
                log.info(f"  ✅ Successfully processed {student.name}")
            else:
                result["status"] = "submit_failed"
                result["error"] = "Could not click submit button"

            # Step 7: Navigate back
            self.go_back_to_batch_list()

        except Exception as e:
            result["status"] = "error"
            result["error"] = str(e)
            log.error(f"  ✗ Error processing {student.name}: {e}")
            try:
                self.go_back_to_batch_list()
            except Exception:
                pass

        return result

    def run(self, students: list[StudentData], start_from: int = 1,
            single_student: Optional[str] = None):
        """Run the automation for all students."""

        if single_student:
            students = [s for s in students if s.enrollment_no == single_student]
            if not students:
                log.error(f"Student {single_student} not found in Excel data")
                return

        total = len(students)
        log.info(f"\n{'#'*60}")
        log.info(f"Starting automation for {total} students")
        log.info(f"Dry run: {self.dry_run}")
        log.info(f"Starting from: student #{start_from}")
        log.info(f"{'#'*60}\n")

        # Navigate to batch page
        self.navigate_to_batch()
        self.click_approved_applicants_tab()

        # Process each student
        for student in students:
            if student.serial_no < start_from:
                log.info(f"Skipping student #{student.serial_no} (before start_from={start_from})")
                continue

            result = self.process_student(student, total)
            self.results.append(result)

            done = len([r for r in self.results if r["status"] in ("success", "already_uploaded")])
            failed = len([r for r in self.results if r["status"] not in ("success", "already_uploaded", "pending")])
            log.info(f"Progress: {done} done, {failed} failed, "
                     f"{total - len(self.results)} remaining")

        self.print_summary()

    def print_summary(self):
        """Print a summary of all results."""
        log.info(f"\n{'#'*60}")
        log.info("AUTOMATION COMPLETE - SUMMARY")
        log.info(f"{'#'*60}")

        success = [r for r in self.results if r["status"] == "success"]
        skipped = [r for r in self.results if r["status"] == "already_uploaded"]
        failed = [r for r in self.results if r["status"] not in ("success", "already_uploaded")]

        log.info(f"  ✅ Successfully uploaded: {len(success)}")
        log.info(f"  ⏭️  Already uploaded (skipped): {len(skipped)}")
        log.info(f"  ❌ Failed: {len(failed)}")

        if failed:
            log.info("\nFailed students:")
            for r in failed:
                log.info(f"  - {r['name']} ({r['enrollment']}): {r['status']} - {r['error']}")

        results_file = f"results_{BATCH_ID}.json"
        with open(results_file, "w") as f:
            json.dump(self.results, f, indent=2)
        log.info(f"\nDetailed results saved to {results_file}")


# ─── Helper for Active Page ──────────────────────────────────────────────────

def get_active_page(context) -> Page:
    """Find the page that is on the batch details page, or default to the last one."""
    pages = context.pages
    if not pages:
        raise RuntimeError("No pages open in browser context")
    
    # Try to find the page with the batch details URL or batch ID
    for p in reversed(pages):
        try:
            url = p.url
            if "view-batch-details" in url or "master-assessor" in url or "PENDING" in url or BATCH_ID in url:
                return p
        except Exception:
            pass
            
    # Fallback to the last page
    return pages[-1]


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    global BATCH_ID, EXCEL_FILE, BATCH_PAGE_URL

    parser = argparse.ArgumentParser(description="SIDH Marks Entry Automation")
    parser.add_argument("--excel", type=str, default="Result Sheet _ 3391657, 21-05-2026.xlsx", help="Excel file name to read")
    parser.add_argument("--batch-id", type=str, default="3391657", help="Batch ID on portal")
    parser.add_argument("--dry-run", action="store_true", help="Fill marks but don't submit")
    parser.add_argument("--start-from", type=int, default=1, help="Start from student number N")
    parser.add_argument("--student", type=str, help="Process only this enrollment number")
    parser.add_argument("--slow", action="store_true", help="Add extra delays (for slow connections)")
    args = parser.parse_args()

    # Update configuration dynamically
    BATCH_ID = args.batch_id
    EXCEL_FILE = args.excel
    BATCH_PAGE_URL = f"{BASE_URL}/admin-profile/assessor/master-assessor/view-batch-details-new/PENDING/{BATCH_ID};batches=assessment"

    # Add FileHandler dynamically for this batch ID
    log_file = f"automation_{BATCH_ID}.log"
    file_handler = logging.FileHandler(log_file, encoding="utf-8")
    file_handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    logging.getLogger().addHandler(file_handler)

    if args.slow:
        global DELAY_BETWEEN_FIELDS, DELAY_AFTER_SEARCH, DELAY_AFTER_NAVIGATE
        DELAY_BETWEEN_FIELDS = 0.4
        DELAY_AFTER_SEARCH = 4.0
        DELAY_AFTER_NAVIGATE = 5.0

    # Read Excel data first
    excel_path = Path(EXCEL_FILE)
    if not excel_path.exists():
        log.error(f"Excel file not found: {EXCEL_FILE}")
        sys.exit(1)

    students = read_excel_data(str(excel_path))

    # Launch browser - user will log in manually in the SAME session
    print()
    print("=" * 60)
    print("  SIDH Marks Entry Automation")
    print("=" * 60)
    print()
    print("A browser will open now. Please:")
    print("  1. Log in to the portal as PMU")
    print("  2. Make sure you're fully logged in")
    print("  3. Come back here and press ENTER")
    print()
    print(f"Students to process: {len(students)}")
    print(f"Dry run: {args.dry_run}")
    print()

    user_data_dir = Path(f"./.playwright_session_{BATCH_ID}").resolve()

    with sync_playwright() as p:
        log.info(f"Launching persistent browser context using profile at: {user_data_dir}")
        context = p.chromium.launch_persistent_context(
            user_data_dir=str(user_data_dir),
            headless=False,
            slow_mo=50,
            viewport={"width": 1366, "height": 768},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                       "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        )

        page = context.pages[0] if context.pages else context.new_page()

        # Try to navigate directly to the batch page
        log.info(f"Navigating to batch details page: {BATCH_PAGE_URL}")
        try:
            page.goto(BATCH_PAGE_URL, wait_until="domcontentloaded")
            time.sleep(2)
        except Exception as e:
            log.warning(f"Initial navigation failed/redirected: {e}")

        # Check if we are logged in (e.g. title is correct, we're not redirected to a login page,
        # or we see the batch details header/applicants tab)
        is_logged_in = False
        try:
            # Check all open pages in context
            for pge in context.pages:
                if pge.locator(f"text=Batch ID - {BATCH_ID}").first.is_visible(timeout=1000) or \
                   pge.locator("text=Approved Applicants").first.is_visible(timeout=1000) or \
                   pge.locator("text=Approved Applicant").first.is_visible(timeout=1000):
                    is_logged_in = True
                    break
        except Exception:
            pass

        if not is_logged_in:
            print()
            print("=" * 60)
            print("  🔒 Session expired or not logged in")
            print("=" * 60)
            print("Please log in manually on the browser window.")
            print(f"And make sure you are navigated to: {BATCH_PAGE_URL}")
            print()
            input(">>> Press ENTER after you've logged in successfully... ")
            print()
        else:
            print()
            print("✅ Session active! Already logged in.")
            print()

        # Find the active page where the batch details are loaded (handles multi-tab login redirects)
        active_page = get_active_page(context)
        log.info(f"Targeting active page: {active_page.url}")

        print("🚀 Starting automation... Watch the browser!")
        print()

        # Run automation in the SAME browser session (no session expiry!)
        automation = SIDHAutomation(page=active_page, dry_run=args.dry_run)
        automation.run(
            students=students,
            start_from=args.start_from,
            single_student=args.student,
        )

        print()
        print("Browser will stay open for 30 seconds for you to review...")
        print("(Close it manually or wait)")
        time.sleep(30)

        context.close()

    log.info("Done!")


if __name__ == "__main__":
    main()
